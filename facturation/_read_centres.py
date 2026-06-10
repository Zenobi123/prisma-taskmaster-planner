import openpyxl, json, sys, os
sys.stdout.reconfigure(encoding='utf-8')

files = [
    r'C:\Users\HP\Downloads\centres_impots_region_centre_cameroun.xlsx',
    r'C:\Users\HP\Downloads\centres_impots_autres_regions_cameroun.xlsx'
]

data = {}
for f in files:
    wb = openpyxl.load_workbook(f, data_only=True)
    fname = os.path.basename(f)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = [row for row in ws.iter_rows(values_only=True)]
        key = fname + ' :: ' + sheet
        data[key] = rows

with open('_centres_impots_data.json', 'w', encoding='utf-8') as out:
    json.dump(data, out, ensure_ascii=False, indent=2, default=str)
print('Saved')
